#[terminal]$: pip install openpyxl
from pathlib import Path
from openpyxl import Workbook
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

#Defining directories
ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xls'

workbook = Workbook()
worksheet : Worksheet = workbook.active

#Creating the headers
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Note')

workbook.save(WORKBOOK_PATH)

students = [
    #name #age #note
    ['John', 14, 5.5],
    ['Marie', 13, 9.7],
    ['Lutz', 15, 8.8],
    ['Carl', 16, 10],
]

for i, student_row in enumerate(students, start=2):
    # print(i, student_row) #AAA
    for j, student_column in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_column)

# for student in students:
#     worksheet.append(student) # Mesmo resultado do for acima.

workbook.save(WORKBOOK_PATH)

# Edson Copque | https://linktr.ee/edsoncopque

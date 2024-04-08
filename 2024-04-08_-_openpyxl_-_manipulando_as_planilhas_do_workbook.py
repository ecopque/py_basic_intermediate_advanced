#[terminal]$: pip install openpyxl
from pathlib import Path #1:
from openpyxl import Workbook #2:
import openpyxl #3:
from openpyxl.worksheet.worksheet import Worksheet #4:

ROOT_FOLDER = Path(__file__).parent #5:
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xls' #6:

workbook = Workbook() #7:
# worksheet : Worksheet = workbook.active #8:
sheetname = 'LOG' #18:
workbook.create_sheet(sheetname, 0) #19:
worksheet : Worksheet = workbook[sheetname]
print(workbook.sheetnames) #20:
workbook.remove(workbook['Sheet']) #21:
print(workbook.sheetnames) #22:

worksheet.cell(1, 1, 'Name') #9:
worksheet.cell(1, 2, 'Age') #9:
worksheet.cell(1, 3, 'Note') #9:

workbook.save(WORKBOOK_PATH) #10:

students = [
    #name #age #note
    ['John', 14, 5.5],
    ['Marie', 13, 9.7],
    ['Lutz', 15, 8.8],
    ['Carl', 16, 10],
] #11:

for i, student_row in enumerate(students, start=2): #12:
    # print(i, student_row) #18:
    for j, student_column in enumerate(student_row, start=1): #13:
        worksheet.cell(i, j, student_column) #14: #17:

# for student in students:
#     worksheet.append(student) #15:

workbook.save(WORKBOOK_PATH) #16:

#Explanation of openpyxl: It is an open-source library used to work with Excel files (.xlsx). It allows reading, writing, and manipulating Excel spreadsheets programmatically, which is useful for automating tasks related to creating and processing spreadsheets in Python.
#1: Imports the Path class from the pathlib module, which is used to handle file and directory paths independently of the platform.
#2: Imports the Workbook class from the openpyxl module, which is a Python library for manipulating Excel files.
#3: Imports the openpyxl module, which contains functionalities for dealing with Excel files.
#4: Imports the Worksheet class from the openpyxl.worksheet.worksheet module, which is used to represent a worksheet within an Excel file.
#5: Defines the constant ROOT_FOLDER as the parent directory of the current executing file.
#6: Defines the path to the spreadsheet file as being in the ROOT_FOLDER directory with the file name 'workbook.xls'.
#7: Creates a new Workbook object, which is an Excel workbook.
#8: Gets the active worksheet from the Workbook object and assigns it to the worksheet variable, which is typed as an instance of Worksheet.
#9: Line x, column y, cell value z.
#10: Saves the Excel workbook at the specified path in WORKBOOK_PATH.
#11: Lists students containing information about the students (name, age, and grade).
#12: Enumerates the elements of the students list starting from 2, i.e., from the second row of the spreadsheet.
#13: Enumerates the elements of each student's row, starting from 1.
#14: Sets the values of the elements of the student's row in the spreadsheet.
#15: We will get the same result as the "for" loop above/previous.
#16: Saves the Excel workbook again after filling in the students' data.
#17: Answer: see image.

#18: This line defines a variable named sheetname and assigns the string 'LOG' to it. This variable will be used later to create a sheet with the name 'LOG' in the workbook.
#19: This line creates a new sheet in the workbook and assigns it the name stored in the sheetname variable (which is 'LOG' in this case). The second argument (0) specifies the position of the sheet within the workbook, where 0 indicates the first position.
#20: This line prints the list of sheet names currently present in the workbook.  After creating the 'LOG' sheet with the previous line, this will likely print ['Sheet', 'LOG'].
#21: This line removes the sheet named 'Sheet' from the workbook. Since the script created a new sheet named 'LOG' earlier, removing the default 'Sheet' helps with organization.
#22: This line again prints the list of sheet names after removing the 'Sheet'. This should now print only ['LOG'], confirming the removal.